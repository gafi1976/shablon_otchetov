# excel_handler.py
# Один файл shablon.xlsx с двумя листами:
#   Лист 1: "Spisaniye"  — данные для акта СПИСАНИЯ
#   Лист 2: "Ustanovka"  — данные для акта УСТАНОВКИ
#
# ════════════════════════════════════════════════════
#  ЛИСТ "Spisaniye" — структура (строка 1 = заголовки)
# ════════════════════════════════════════════════════
#  A  Tashkilot       — Название организации
#  B  Rahbar          — Руководитель (ФИО)
#  C  Muhandis1       — Инженер 1 (ФИО)
#  D  Muhandis2       — Инженер 2 (ФИО)
#  E  Inventar        — Инвентарный номер  (ключ группировки)
#  F  Qurilma         — Название оборудования
#  G  Qism            — Компонент
#  H  Holat           — Состояние: Yaroqli / Yaroqsiz
#  I  Sabab           — Причина / неисправность
#
# ════════════════════════════════════════════════════
#  ЛИСТ "Ustanovka" — структура (строка 1 = заголовки)
# ════════════════════════════════════════════════════
#  A  Tashkilot       — Название организации
#  B  Manzil          — Адрес организации
#  C  Rahbar          — Руководитель (ФИО)
#  D  Muhandis1       — Инженер 1 (ФИО)
#  E  Lavozim1        — Должность инженера 1
#  F  Muhandis2       — Инженер 2 (ФИО)
#  G  Lavozim2        — Должность инженера 2
#  H  Sana            — Дата установки (дд.мм.гггг)
#  I  Uskuna          — Название / модель оборудования
#  J  Seriya          — Серийный номер
#  K  Joyi            — Место установки

import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime, date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation


# ─────────────────────────────────────────────────
#  Стили
# ─────────────────────────────────────────────────

_FONT_HDR  = Font(name='Times New Roman', bold=True, color='FFFFFF', size=11)
_FONT_DATA = Font(name='Times New Roman', size=10)
_FILL_S    = PatternFill('solid', fgColor='2E4057')   # списание — тёмно-синий
_FILL_U    = PatternFill('solid', fgColor='1A5276')   # установка — синий
_FILL_ES   = PatternFill('solid', fgColor='EBF5FB')
_FILL_ODD  = PatternFill('solid', fgColor='FFFFFF')
_ALIGN_C   = Alignment(horizontal='center', vertical='center', wrap_text=True)
_ALIGN_L   = Alignment(horizontal='left',   vertical='center', wrap_text=True)


def _border():
    s = Side(style='thin', color='AAAAAA')
    return Border(left=s, right=s, top=s, bottom=s)


def _make_sheet(ws, title_text, title_color, headers_widths, hdr_fill):
    """
    Создаёт лист с:
      Строка 1 — цветная строка-заголовок листа (название)
      Строка 2 — заголовки колонок
      Строки 3+ — пустые строки для данных
    """
    ws.sheet_view.showGridLines = False

    # Строка 1 — заголовок листа
    ncols = len(headers_widths)
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=ncols)
    cell = ws.cell(row=1, column=1, value=title_text)
    cell.font      = Font(name='Times New Roman', bold=True, size=14, color='FFFFFF')
    cell.fill      = PatternFill('solid', fgColor=title_color)
    cell.alignment = _ALIGN_C
    ws.row_dimensions[1].height = 28

    # Строка 2 — заголовки колонок
    ws.row_dimensions[2].height = 38
    for ci, (header, width) in enumerate(headers_widths, start=1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(ci)].width = width
        cell = ws.cell(row=2, column=ci, value=header)
        cell.font      = _FONT_HDR
        cell.fill      = hdr_fill
        cell.alignment = _ALIGN_C
        cell.border    = _border()

    # Строки 3–52 — пустые для ввода данных (50 строк)
    # Для Spisaniye: числовые колонки E(5), H(8), J-O(10-15), P-T(16-20)
    # Для Ustanovka: колонка H(8)
    center_cols = {5, 8, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20}
    for ri in range(3, 53):
        ws.row_dimensions[ri].height = 20
        fill = _FILL_ES if ri % 2 == 0 else _FILL_ODD
        for ci in range(1, ncols + 1):
            cell = ws.cell(row=ri, column=ci)
            cell.font      = _FONT_DATA
            cell.fill      = fill
            cell.border    = _border()
            cell.alignment = _ALIGN_C if ci in center_cols else _ALIGN_L

    ws.freeze_panes = 'A3'


# ─────────────────────────────────────────────────
#  СОЗДАНИЕ shablon.xlsx
# ─────────────────────────────────────────────────

SHABLON_NAME = 'shablon.xlsx'

# ════════════════════════════════════════════════════════════════
#  ШАБЛОН СПИСАНИЯ — структура колонок (строка 1=title, строка 2=заголовки)
# ════════════════════════════════════════════════════════════════
#  A  Tashkilot        — Название организации
#  B  Direktor         — Директор организации (ФИО)
#  C  Boshliq          — Начальник отдела (ФИО)
#  D  Buyruq           — Номер приказа комиссии
#  E  Inventar         — Инвентарный номер  (ключ группировки)
#  F  Qurilma          — Название оборудования
#  G  Qism             — Компонент
#  H  Holat            — Состояние (Yaroqli / Yaroqsiz)
#  I  Sabab            — Причина / неисправность
#  J  Narxi            — Стоимость (сўм)
#  K  Amort            — Амортизация чегирмалари (сўм)
#  L  Norma            — Норма амортизации (%)
#  M  Qoldiq           — Қолдиқ нархи (сўм)
#  N  Debet            — Дебет счёт (напр. 9210)
#  O  Kredit           — Кредит счёт (напр. 0150)
#  P  Ish_chiq_yil     — Год производства (изготовления)
#  Q  Kelgan_yil       — Год прихода (поступления)
#  R  Ishga_yil        — Год ввода в эксплуатацию
#  S  Tamir_soni       — Количество ремонтов
#  T  Tamir_summa      — Сумма затрат на ремонты (сўм)
#  U  Bulim            — Подразделение / отдел
#  V  Azо1_fio         — Член комиссии 1: ФИО
#  W  Azо1_lavozim     — Член комиссии 1: должность
#  X  Azо2_fio         — Член комиссии 2: ФИО
#  Y  Azо2_lavozim     — Член комиссии 2: должность
#  Z  Azо3_fio         — Член комиссии 3: ФИО
#  AA Azо3_lavozim     — Член комиссии 3: должность
#  AB Azо4_fio         — Член комиссии 4: ФИО
#  AC Azо4_lavozim     — Член комиссии 4: должность
#  AD Azо5_fio         — Член комиссии 5: ФИО
#  AE Azо5_lavozim     — Член комиссии 5: должность

HEADERS_S = [
    # Основные
    ('Tashkilot\n(Организация)',                  28),  # A
    ('Direktor\n(Директор организации)',          24),  # B
    ('Boshliq\n(Начальник отдела)',               24),  # C
    ('Buyruq raqami\n(Номер приказа)',             18),  # D
    ('Inventar ★\n(Инв. номер)',                  16),  # E
    ('Qurilma\n(Оборудование)',                   26),  # F
    ('Qism\n(Компонент)',                          22),  # G
    ('Holat\n(Yaroqli / Yaroqsiz)',               16),  # H
    ('Sabab\n(Причина)',                           24),  # I
    # Финансовые
    ('Narxi\n(Стоимость, сўм)',                   16),  # J
    ('Amort\n(Амортизация, сўм)',                 18),  # K
    ('Norma\n(Норма, %)',                          12),  # L
    ('Qoldiq\n(Қолдиқ, сўм)',                    16),  # M
    ('Debet\n(Дебет счёт)',                        14),  # N
    ('Kredit\n(Кредит счёт)',                      14),  # O
    # История
    ("Ishlab chiqarilgan yil\n(Год производства)", 18), # P
    ("Kelgan yil\n(Год прихода)",                  14), # Q
    ("Ishga tushirilgan yil\n(Год ввода)",         14), # R
    ("Ta'mirlash soni\n(Кол-во ремонтов)",        14), # S
    ("Ta'mirlash summasi\n(Сумма ремонтов, сўм)", 18), # T
    ("Bo'lim\n(Подразделение)",                   22), # U
    # Члены комиссии (динамически — до 5 пар ФИО + должность)
    ('Azо 1 FIO\n(Чл.ком. 1 ФИО)',               22),  # V
    ('Azо 1 lavozim\n(Чл.ком. 1 должность)',      20),  # W
    ('Azо 2 FIO\n(Чл.ком. 2 ФИО)',               22),  # X
    ('Azо 2 lavozim\n(Чл.ком. 2 должность)',      20),  # Y
    ('Azо 3 FIO\n(Чл.ком. 3 ФИО)',               22),  # Z
    ('Azо 3 lavozim\n(Чл.ком. 3 должность)',      20),  # AA
    ('Azо 4 FIO\n(Чл.ком. 4 ФИО)',               22),  # AB
    ('Azо 4 lavozim\n(Чл.ком. 4 должность)',      20),  # AC
    ('Azо 5 FIO\n(Чл.ком. 5 ФИО)',               22),  # AD
    ('Azо 5 lavozim\n(Чл.ком. 5 должность)',      20),  # AE
]

# Заголовки листа Ustanovka (11 колонок)
HEADERS_U = [
    ('Tashkilot\n(Организация)',           28),
    ('Manzil\n(Адрес)',                    22),
    ('Rahbar\n(Руководитель)',             22),
    ('Muhandis1\n(Инженер 1)',             22),
    ('Lavozim1\n(Должность 1)',            20),
    ('Muhandis2\n(Инженер 2)',             22),
    ('Lavozim2\n(Должность 2)',            20),
    ('Sana\n(Дата дд.мм.гггг)',            16),
    ('Uskuna\n(Оборудование / Модель)',    26),
    ('Seriya\n(Серийный номер)',           18),
    ("Joyi\n(Место установки)",           22),
]


def create_template(path: str) -> str:
    """
    Создаёт shablon.xlsx с двумя листами:
      Лист "Spisaniye" — для акта СПИСАНИЯ
      Лист "Ustanovka" — для акта УСТАНОВКИ
    Только заголовки, пустые строки для заполнения.
    """
    wb = Workbook()

    # Лист 1 — Spisaniye
    ws_s = wb.active
    ws_s.title = 'Spisaniye'
    _make_sheet(ws_s,
                'АКТ СПИСАНИЯ — заполните данные начиная с строки 3',
                '2E4057', HEADERS_S, _FILL_S)

    # Валидация колонки H (Holat)
    dv = DataValidation(type='list',
                        formula1='"Yaroqli,Yaroqsiz"',
                        allow_blank=True, showDropDown=False)
    dv.sqref = 'H3:H200'
    ws_s.add_data_validation(dv)

    # Лист 2 — Ustanovka
    ws_u = wb.create_sheet(title='Ustanovka')
    _make_sheet(ws_u,
                'АКТ УСТАНОВКИ — заполните данные начиная с строки 3',
                '1A5276', HEADERS_U, _FILL_U)

    wb.save(path)
    return path


# ─────────────────────────────────────────────────
#  Для обратной совместимости (вызывается из app.py)
# ─────────────────────────────────────────────────

def create_spisan_template(path: str) -> str:
    """Создаёт полный shablon.xlsx (оба листа) в папке path."""
    import os
    folder = os.path.dirname(path)
    out    = os.path.join(folder, SHABLON_NAME)
    return create_template(out)


def create_ust_template(path: str) -> str:
    """Создаёт полный shablon.xlsx (оба листа) в папке path."""
    import os
    folder = os.path.dirname(path)
    out    = os.path.join(folder, SHABLON_NAME)
    return create_template(out)


# ─────────────────────────────────────────────────
#  Низкоуровневый читатель xlsx
# ─────────────────────────────────────────────────

def _read_sheet(path: str, sheet_index: int = 0) -> list:
    """
    Читает указанный лист xlsx-файла.
    Возвращает list строк; строка 0 = строка 1 Excel.
    Каждая строка — dict {col_index: value} ИЛИ list с правильными позициями.
    Пустые ячейки = '' на правильном месте (не сжимает).
    """
    with zipfile.ZipFile(path) as z:
        names = z.namelist()

        # Shared strings
        strings = []
        if 'xl/sharedStrings.xml' in names:
            ns = {'n': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            for si in ET.parse(z.open('xl/sharedStrings.xml')).findall('.//n:si', ns):
                strings.append(''.join(t.text or '' for t in si.findall('.//n:t', ns)))

        # Все листы
        sheets = sorted(f for f in names if f.startswith('xl/worksheets/sheet'))
        if sheet_index >= len(sheets):
            return []

        ns = {'n': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        all_rows = ET.parse(z.open(sheets[sheet_index])).findall('.//n:row', ns)
        if not all_rows:
            return []

        def col_letter_to_num(ref: str) -> int:
            """A→1, B→2, AA→27 и т.д."""
            col = ''.join(c for c in ref if c.isalpha())
            n = 0
            for ch in col.upper():
                n = n * 26 + (ord(ch) - ord('A') + 1)
            return n - 1   # 0-based

        max_row = max(int(r.get('r', 0)) for r in all_rows)
        # Находим максимальную колонку по всему листу
        max_col = 0
        for row_el in all_rows:
            for c in row_el.findall('n:c', ns):
                ci = col_letter_to_num(c.get('r', 'A1'))
                if ci > max_col:
                    max_col = ci

        result = [[] for _ in range(max_row + 1)]

        for row_el in all_rows:
            rn = int(row_el.get('r', 0))
            row = [''] * (max_col + 1)   # правильный размер с пустыми
            for c in row_el.findall('n:c', ns):
                ci = col_letter_to_num(c.get('r', 'A1'))
                t  = c.get('t', '')
                v  = c.find('n:v', ns)
                if v is not None and v.text is not None:
                    row[ci] = strings[int(v.text)] if t == 's' else v.text
                else:
                    row[ci] = ''
            result[rn] = row

        return result[1:]   # индекс 0 пустой


def _g(row: list, idx: int) -> str:
    """Безопасно возвращает ячейку как строку, убирает переносы строк."""
    if idx >= len(row):
        return ''
    v = row[idx]
    if not v:
        return ''
    # Убираем переносы строк из значений ячеек
    return str(v).strip().replace('\n', ' ').replace('\r', '').replace('\t', ' ')


def _to_date(val: str) -> str:
    """Excel serial number → дд.мм.гггг; строку оставляет как есть."""
    if not val:
        return ''
    val = val.strip()
    # Уже в нужном формате
    if len(val) == 10 and val[2] == '.' and val[5] == '.':
        return val
    # Excel serial number
    try:
        n = float(val)
        if n > 1000:
            return (date(1899, 12, 30) + timedelta(days=int(n))).strftime('%d.%m.%Y')
    except ValueError:
        pass
    return val


def _get_sheet_index(path: str, sheet_name: str) -> int:
    """Возвращает индекс листа по его имени."""
    try:
        with zipfile.ZipFile(path) as z:
            tree = ET.parse(z.open('xl/workbook.xml'))
            # Пробуем оба варианта namespace
            for ns_uri in [
                'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
                '',
            ]:
                ns = {'n': ns_uri} if ns_uri else {}
                finder = './/n:sheet' if ns_uri else './/sheet'
                sheets = tree.findall(finder, ns) if ns_uri else tree.findall(finder)
                for i, s in enumerate(sheets):
                    name = s.get('name', '')
                    if name.lower() == sheet_name.lower():
                        return i
    except Exception:
        pass
    return -1   # не найден


def _detect_format(rows: list) -> str:
    """
    Определяет формат файла.
    Новый shablon.xlsx: строка 1 = title (АКТ...), строка 2 = заголовки колонок.
    Старые файлы: строка 1 = заголовки колонок, данные с строки 2.
    """
    if not rows:
        return 'unknown'

    # Проверяем строку 0 (= строка 1 Excel) на наличие title нового шаблона
    row0_text = ' '.join(_g(rows[0], i) for i in range(min(3, len(rows[0])))).lower()
    is_new = any(w in row0_text for w in ('акт ', 'шаблон', 'akt ', 'shablon'))

    # Для нового шаблона заголовки в строке 1 (index 1), для старого — в строке 0
    hdr_row  = rows[1] if (is_new and len(rows) > 1) else rows[0]
    hdr      = [_g(hdr_row, i).lower() for i in range(min(15, len(hdr_row)))]
    joined   = ' '.join(hdr)

    if is_new:
        if 'inventar' in joined and ('qurilma' in joined or 'qism' in joined):
            return 'new_spisan'
        if 'uskuna' in joined or 'joyi' in joined or 'manzil' in joined:
            return 'new_ust'

    # Старые форматы
    if 'qurilma nomi' in joined or (hdr and 'qurilma' in hdr[0]):
        return 'old_spisan'
    if hdr and ('data' in hdr[0] or '{num_otch}' in joined or 'installation' in joined):
        return 'old_ust'

    return 'unknown'


# ─────────────────────────────────────────────────
#  ЧТЕНИЕ ЛИСТА "Spisaniye"  /  автоопределение
# ─────────────────────────────────────────────────

def read_spisan_excel(path: str) -> list:
    """
    Читает данные для акта СПИСАНИЯ из xlsx-файла.
    Поддерживает три формата автоматически:

    1. НОВЫЙ (shablon.xlsx, лист 'Spisaniye'):
       Строка 1 = заголовки, данные с строки 2.
       A=Tashkilot, B=Rahbar, C=Muhandis1, D=Muhandis2,
       E=Inventar★, F=Qurilma, G=Qism, H=Holat, I=Sabab

    2. СТАРЫЙ (shablon_spisan.xlsx, лист 'Shablon'):
       Строка 1 = заголовки, данные с строки 2.
       A=Qurilma nomi, B=Qismlar nomi, C=Foydalanishga yaroqliligi,
       D=Nosozlik belgilari, E=inv_num, F=Tashilot nomi,
       G=boss name, H=enginer1, I=enginer2

    Возвращает список групп по инвентарному номеру.
    """
    # Сначала пробуем найти лист 'Spisaniye'
    idx = _get_sheet_index(path, 'Spisaniye')
    if idx < 0:
        idx = 0   # берём первый лист

    rows = _read_sheet(path, idx)
    if not rows:
        raise ValueError(f'Файл пуст или не читается: {path}')

    fmt = _detect_format(rows)

    if fmt == 'new_spisan':
        # Строка 1 = title листа, строка 2 = заголовки колонок, данные с строки 3
        return _parse_spisan_new(rows[2:])
    elif fmt == 'old_spisan':
        # Строка 1 = заголовки, данные с строки 2
        return _parse_spisan_old(rows[1:])
    else:
        # Неизвестный — пробуем определить по строке 1
        # Если строка 1 содержит цветной title (нет данных) — данные с строки 3
        row0_vals = [_g(rows[0], i) for i in range(min(5, len(rows[0])))]
        if all(not v or 'АКТ' in v or 'шаблон' in v.lower() for v in row0_vals if v):
            return _parse_spisan_new(rows[2:])
        return _parse_spisan_new(rows[1:])


def _parse_spisan_new(data_rows: list) -> list:
    """
    Парсит данные нового формата (shablon.xlsx, лист Spisaniye):
    A=Tashkilot, B=Direktor, C=Boshliq, D=Buyruq, E=Inventar★,
    F=Qurilma, G=Qism, H=Holat, I=Sabab,
    J=Narxi, K=Amort, L=Norma, M=Qoldiq, N=Debet, O=Kredit,
    P=Ish_chiq_yil, Q=Kelgan_yil, R=Ishga_yil,
    S=Tamir_soni, T=Tamir_summa, U=Bulim,
    V=Azo1_fio, W=Azo1_lavozim, X=Azo2_fio, Y=Azo2_lavozim,
    Z=Azo3_fio, AA=Azo3_lavozim, AB=Azo4_fio, AC=Azo4_lavozim,
    AD=Azo5_fio, AE=Azo5_lavozim
    """
    inv_data  = {}
    inv_order = []

    for row in data_rows:
        if not row or not any(row):
            continue

        tashkilot   = _g(row,  0)   # A
        direktor    = _g(row,  1)   # B
        boshliq     = _g(row,  2)   # C
        buyruq      = _g(row,  3)   # D
        inventar    = _g(row,  4)   # E  ← ключ
        qurilma     = _g(row,  5)   # F
        qism        = _g(row,  6)   # G
        holat       = _g(row,  7)   # H
        sabab       = _g(row,  8)   # I
        narxi       = _g(row,  9)   # J
        amort       = _g(row, 10)   # K
        norma       = _g(row, 11)   # L
        qoldiq      = _g(row, 12)   # M
        debet       = _g(row, 13)   # N
        kredit      = _g(row, 14)   # O
        ish_chiq    = _g(row, 15)   # P
        kelgan      = _g(row, 16)   # Q
        ishga       = _g(row, 17)   # R
        tamir_soni  = _g(row, 18)   # S
        tamir_summa = _g(row, 19)   # T
        bulim       = _g(row, 20)   # U

        # Члены комиссии — динамически (до 5 пар)
        azolar = []
        for i in range(5):
            fio     = _g(row, 21 + i*2)     # V, X, Z, AB, AD
            lavozim = _g(row, 21 + i*2 + 1) # W, Y, AA, AC, AE
            if fio:
                azolar.append({'fio': fio, 'lavozim': lavozim})

        if not inventar:
            continue

        if inventar not in inv_data:
            inv_data[inventar] = {
                'inv_number':   inventar,
                'org_name':     tashkilot,
                'direktor':     direktor,
                'boshliq':      boshliq,
                'buyruq':       buyruq,
                'region':       '',
                'doc_date':     datetime.now().strftime('%d.%m.%Y'),
                'narxi':        narxi,
                'amort':        amort,
                'norma':        norma,
                'qoldiq':       qoldiq,
                'debet':        debet or '9210',
                'kredit':       kredit or '0150',
                'ish_chiq_yil': ish_chiq,
                'kelgan_yil':   kelgan,
                'ishga_yil':    ishga,
                'tamir_soni':   tamir_soni,
                'tamir_summa':  tamir_summa,
                'bulim':        bulim,
                'azolar':       azolar,
                'devices':      {},
                'dev_order':    [],
            }
            inv_order.append(inventar)

        g = inv_data[inventar]
        # Обновляем если пустые
        for key, val in [
            ('org_name', tashkilot), ('direktor', direktor),
            ('boshliq', boshliq),    ('buyruq', buyruq),
            ('narxi', narxi),        ('amort', amort),
            ('norma', norma),        ('qoldiq', qoldiq),
            ('ish_chiq_yil', ish_chiq), ('kelgan_yil', kelgan),
            ('ishga_yil', ishga),    ('tamir_soni', tamir_soni),
            ('tamir_summa', tamir_summa), ('bulim', bulim),
        ]:
            if not g[key] and val:
                g[key] = val
        if not g['azolar'] and azolar:
            g['azolar'] = azolar

        if qurilma:
            if qurilma not in g['devices']:
                g['devices'][qurilma] = []
                g['dev_order'].append(qurilma)
            if qism:
                g['devices'][qurilma].append({
                    'part_name': qism,
                    'condition': holat,
                    'defect':    sabab,
                })

    return _build_groups(inv_data, inv_order)


def _parse_spisan_old(data_rows: list) -> list:
    """
    Парсит данные старого формата (shablon_spisan.xlsx):
    A=Qurilma nomi, B=Qismlar nomi, C=Foydalanishga yaroqliligi,
    D=Nosozlik belgilari, E=inv_num, F=Tashilot nomi,
    G=boss name, H=enginer1, I=enginer2
    """
    inv_data  = {}
    inv_order = []

    for row in data_rows:
        if not row or not any(row):
            continue
        qurilma   = _g(row, 0)
        qism      = _g(row, 1)
        holat     = _g(row, 2)
        sabab     = _g(row, 3)
        inventar  = _g(row, 4)
        tashkilot = _g(row, 5)
        rahbar    = _g(row, 6)
        muh1      = _g(row, 7)
        muh2      = _g(row, 8)

        if not inventar:
            continue

        if inventar not in inv_data:
            inv_data[inventar] = {
                'inv_number':      inventar,
                'org_name':        tashkilot,
                'direktor':        '',
                'boshliq':         rahbar,
                'buyruq':          '',
                'region':          '',
                'doc_date':        datetime.now().strftime('%d.%m.%Y'),
                'narxi':           '',
                'amort':           '0.00',
                'norma':           '',
                'qoldiq':          '0.00',
                'debet':           '9210',
                'kredit':          '0150',
                'ish_chiq_yil':    '',
                'kelgan_yil':      '',
                'ishga_yil':       '',
                'tamir_soni':      '',
                'tamir_summa':     '',
                'bulim':           '',
                'azolar':          [
                    {'fio': muh1, 'lavozim': ''} if muh1 else None,
                    {'fio': muh2, 'lavozim': ''} if muh2 else None,
                ],
                'commission_head': rahbar,
                'member1':         muh1,
                'member2':         muh2,
                'devices':         {},
                'dev_order':       [],
            }
            # Убираем None из azolar
            inv_data[inventar]['azolar'] = [a for a in inv_data[inventar]['azolar'] if a]
            inv_order.append(inventar)

        g = inv_data[inventar]
        if not g['org_name']        and tashkilot: g['org_name']        = tashkilot
        if not g['commission_head'] and rahbar:    g['commission_head'] = rahbar
        if not g['member1']         and muh1:      g['member1']         = muh1
        if not g['member2']         and muh2:      g['member2']         = muh2

        if qurilma:
            if qurilma not in g['devices']:
                g['devices'][qurilma] = []
                g['dev_order'].append(qurilma)
            if qism:
                g['devices'][qurilma].append({
                    'part_name': qism,
                    'condition': holat,
                    'defect':    sabab,
                })

    return _build_groups(inv_data, inv_order)


def _build_groups(inv_data: dict, inv_order: list) -> list:
    result = []
    for inv in inv_order:
        g = inv_data[inv]
        result.append({
            'inv_number':   g['inv_number'],
            'org_name':     g.get('org_name', ''),
            'direktor':     g.get('direktor', ''),
            'boshliq':      g.get('boshliq', ''),
            'buyruq':       g.get('buyruq', ''),
            'region':       '',
            'doc_date':     g.get('doc_date', datetime.now().strftime('%d.%m.%Y')),
            # финансовые
            'narxi':        g.get('narxi', ''),
            'amort':        g.get('amort', '0.00'),
            'norma':        g.get('norma', ''),
            'qoldiq':       g.get('qoldiq', '0.00'),
            'debet':        g.get('debet', '9210'),
            'kredit':       g.get('kredit', '0150'),
            # история
            'ish_chiq_yil': g.get('ish_chiq_yil', ''),
            'kelgan_yil':   g.get('kelgan_yil', ''),
            'ishga_yil':    g.get('ishga_yil', ''),
            'tamir_soni':   g.get('tamir_soni', ''),
            'tamir_summa':  g.get('tamir_summa', ''),
            'bulim':        g.get('bulim', ''),
            # члены комиссии — список {'fio': str, 'lavozim': str}
            'azolar':       g.get('azolar', []),
            # обратная совместимость
            'commission_head': g.get('boshliq', ''),
            'member1':         g['azolar'][0]['fio'] if g.get('azolar') else '',
            'member2':         g['azolar'][1]['fio'] if len(g.get('azolar', [])) > 1 else '',
            'devices': [
                {'name': dn, 'parts': g['devices'][dn]}
                for dn in g['dev_order']
            ],
        })
    return result


# ─────────────────────────────────────────────────
#  ЧТЕНИЕ ЛИСТА "Ustanovka"  /  автоопределение
# ─────────────────────────────────────────────────

def read_ust_excel(path: str) -> dict:
    """
    Читает данные для акта УСТАНОВКИ из xlsx-файла.
    Поддерживает два формата автоматически:

    1. НОВЫЙ (shablon.xlsx, лист 'Ustanovka'):
       A=Tashkilot, B=Manzil, C=Rahbar, D=Muhandis1, E=Lavozim1,
       F=Muhandis2, G=Lavozim2, H=Sana, I=Uskuna, J=Seriya, K=Joyi

    2. СТАРЫЙ (shablom_ust.xlsx):
       A=Data, B=num_otch, C=installation date, D=oborud name,
       E=serial num, F=where oborud, G=Organizasiya, H=adress,
       I=пусто, J=boss name, K=engine1, L=job title1,
       M=engine2, N=job title2
    """
    # Сначала пробуем найти лист 'Ustanovka'
    idx = _get_sheet_index(path, 'Ustanovka')
    if idx < 0:
        idx = 0

    rows = _read_sheet(path, idx)
    if not rows:
        raise ValueError(f'Файл пуст или не читается: {path}')

    fmt = _detect_format(rows)

    if fmt in ('new_ust', 'new_spisan'):
        # Строка 1 = title листа, строка 2 = заголовки колонок, данные с строки 3
        return _parse_ust_new(rows[2:])
    else:
        # Старый формат: строка 1 = заголовки, данные с строки 2
        return _parse_ust_old(rows[1:])


def _parse_ust_new(data_rows: list) -> dict:
    """
    Новый формат Ustanovka:
    A=Tashkilot, B=Manzil, C=Rahbar, D=Muhandis1, E=Lavozim1,
    F=Muhandis2, G=Lavozim2, H=Sana, I=Uskuna, J=Seriya, K=Joyi
    """
    org=''; addr=''; boss=''; e1=''; j1=''; e2=''; j2=''
    items = []

    for row in data_rows:
        if not row or not any(row):
            continue
        tashkilot = _g(row,  0)
        manzil    = _g(row,  1)
        rahbar    = _g(row,  2)
        muh1      = _g(row,  3)
        lav1      = _g(row,  4)
        muh2      = _g(row,  5)
        lav2      = _g(row,  6)
        sana      = _g(row,  7)
        uskuna    = _g(row,  8)
        seriya    = _g(row,  9)
        joyi      = _g(row, 10)

        if not org  and tashkilot: org  = tashkilot
        if not addr and manzil:    addr = manzil
        if not boss and rahbar:    boss = rahbar
        if not e1   and muh1:      e1 = muh1; j1 = lav1
        if not e2   and muh2:      e2 = muh2; j2 = lav2

        if not uskuna:
            continue

        items.append({
            'num':           str(len(items) + 1),
            'name':          uskuna,
            'model':         uskuna,
            'serial_number': seriya,
            'inv_number':    '',
            'install_date':  _to_date(sana) or datetime.now().strftime('%d.%m.%Y'),
            'location':      joyi,
            'cost':          '',
            'condition':     'Yangi',
            'note':          joyi,
        })

    doc_date = items[0]['install_date'] if items else datetime.now().strftime('%d.%m.%Y')
    return {
        'org_name':        org,
        'region':          org,
        'address':         addr,
        'doc_number':      '1',
        'doc_date':        doc_date,
        'commission_head': boss,
        'member1':         e1,
        'member1_title':   j1 or 'Yetakchi muhandis',
        'member2':         e2,
        'member2_title':   j2 or 'Yetakchi muhandis',
        'items':           items,
    }


def _parse_ust_old(data_rows: list) -> dict:
    """
    Старый формат shablom_ust.xlsx:
    A(0)=Data, B(1)=num, C(2)=inst_date, D(3)=oborud_name,
    E(4)=serial, F(5)=where, G(6)=Org, H(7)=addr,
    I(8)=ПУСТАЯ, J(9)=boss, K(10)=eng1, L(11)=job1,
    M(12)=eng2, N(13)=job2
    """
    org=''; addr=''; boss=''; e1=''; j1=''; e2=''; j2=''
    doc_date = ''
    items = []

    for row in data_rows:
        if not row or not any(row):
            continue
        raw_doc   = _g(row,  0)   # A  дата акта
        num       = _g(row,  1)   # B  номер строки
        raw_inst  = _g(row,  2)   # C  дата установки
        uskuna    = _g(row,  3)   # D  оборудование
        seriya    = _g(row,  4)   # E  серийный номер
        where     = _g(row,  5)   # F  место
        org_v     = _g(row,  6)   # G  организация
        addr_v    = _g(row,  7)   # H  адрес
        # col 8 (I) — ПУСТАЯ
        boss_v    = _g(row,  9)   # J  руководитель
        e1_v      = _g(row, 10)   # K  инженер 1
        j1_v      = _g(row, 11)   # L  должность 1
        e2_v      = _g(row, 12)   # M  инженер 2
        j2_v      = _g(row, 13)   # N  должность 2

        if not doc_date and raw_doc: doc_date = _to_date(raw_doc)
        if not org  and org_v:   org  = org_v
        if not addr and addr_v:  addr = addr_v
        if not boss and boss_v:  boss = boss_v
        if not e1   and e1_v:    e1 = e1_v;  j1 = j1_v
        if not e2   and e2_v:    e2 = e2_v;  j2 = j2_v

        if not uskuna:
            continue

        inst = (_to_date(raw_inst) if raw_inst else
                _to_date(raw_doc)  if raw_doc  else
                datetime.now().strftime('%d.%m.%Y'))
        items.append({
            'num':           num or str(len(items) + 1),
            'name':          uskuna,
            'model':         uskuna,
            'serial_number': seriya,
            'inv_number':    '',
            'install_date':  inst,
            'location':      where,
            'cost':          '',
            'condition':     'Yangi',
            'note':          where,
        })

    doc_date = doc_date or (items[0]['install_date'] if items
                            else datetime.now().strftime('%d.%m.%Y'))
    return {
        'org_name':        org,
        'region':          org,
        'address':         addr,
        'doc_number':      '1',
        'doc_date':        doc_date,
        'commission_head': boss,
        'member1':         e1,
        'member1_title':   j1 or 'Yetakchi muhandis',
        'member2':         e2,
        'member2_title':   j2 or 'Yetakchi muhandis',
        'items':           items,
    }
